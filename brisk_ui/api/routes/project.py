"""Project settings API routes.

Reads and writes project configuration from the .brisk/project.json file.
"""

import os
import re
import shutil

import fastapi
import pydantic

from brisk_ui.services.project_config import ProjectConfigService


def sanitize_directory_name(name: str) -> str:
    """Convert a project name to a valid directory name.
    
    - Converts to lowercase
    - Replaces spaces and special chars with hyphens
    - Removes consecutive hyphens
    - Strips leading/trailing hyphens
    """
    # Convert to lowercase and replace spaces/special chars with hyphens
    sanitized = re.sub(r'[^a-zA-Z0-9]+', '-', name.lower())
    # Remove consecutive hyphens
    sanitized = re.sub(r'-+', '-', sanitized)
    # Strip leading/trailing hyphens
    sanitized = sanitized.strip('-')
    return sanitized or "project"

router = fastapi.APIRouter()


class ProjectSettings(pydantic.BaseModel):
    """Project settings model."""
    project_name: str = ""
    project_path: str = ""
    project_description: str = ""
    project_type: str = "classification"  # "classification" or "regression"


class ProjectSettingsUpdate(pydantic.BaseModel):
    """Partial update model for project settings."""
    project_name: str | None = None
    project_path: str | None = None
    project_description: str | None = None
    project_type: str | None = None


@router.get("", response_model=ProjectSettings)
async def get_project_settings(request: fastapi.Request):
    """Get project settings from project.json file."""
    settings = request.app.state.settings
    config_service = ProjectConfigService(settings.project_path)
    
    config_data = config_service.read()
    
    return ProjectSettings(
        project_name=config_data.get("project_name", ""),
        project_path=config_data.get("project_path", str(settings.project_path)),
        project_description=config_data.get("project_description", ""),
        project_type=config_data.get("project_type", "classification"),
    )


class CreateProjectRequest(pydantic.BaseModel):
    """Request model for creating/initializing a project."""
    project_name: str
    project_path: str = ""
    project_description: str = ""
    project_type: str = "classification"  # "classification" or "regression"


class CreateProjectResponse(pydantic.BaseModel):
    """Response model for project creation."""
    project_name: str
    project_path: str
    project_description: str
    project_type: str
    directory_name: str


@router.post("", response_model=CreateProjectResponse)
async def create_project(
    request: fastapi.Request,
    data: CreateProjectRequest,
):
    """Create or initialize project settings.
    
    In create mode (BRISK_UI_CREATE_MODE=true):
    - Creates a new directory named after the sanitized project name
    - The project_path setting is the parent directory
    
    In normal mode:
    - Creates .brisk directory and project.json file in existing project directory
    """
    settings = request.app.state.settings
    create_mode = os.environ.get("BRISK_UI_CREATE_MODE", "false") == "true"
    
    if create_mode:
        # In create mode, use user-provided project_path if given, else settings.project_path
        import pathlib
        parent_dir = pathlib.Path(data.project_path) if data.project_path else settings.project_path
        
        # Validate parent directory exists
        if not parent_dir.exists():
            raise fastapi.HTTPException(
                status_code=400,
                detail=f"Parent directory does not exist: {parent_dir}"
            )
        if not parent_dir.is_dir():
            raise fastapi.HTTPException(
                status_code=400,
                detail=f"Path is not a directory: {parent_dir}"
            )
        
        # Create a subdirectory named after the project
        dir_name = sanitize_directory_name(data.project_name)
        project_dir = parent_dir / dir_name
        
        # Check if directory already exists
        if project_dir.exists():
            raise fastapi.HTTPException(
                status_code=400,
                detail=f"Project directory already exists: {project_dir}"
            )
        
        # Create the project directory and .brisk subdirectory
        brisk_dir = project_dir / ".brisk"
        brisk_dir.mkdir(parents=True, exist_ok=True)
        
        # Use the new project directory for config service
        config_service = ProjectConfigService(project_dir)
        actual_project_path = str(project_dir)
    else:
        # Normal mode - use existing project directory
        dir_name = settings.project_path.name
        brisk_dir = settings.project_path / ".brisk"
        brisk_dir.mkdir(parents=True, exist_ok=True)
        config_service = ProjectConfigService(settings.project_path)
        actual_project_path = data.project_path or str(settings.project_path)
    
    # Write project settings to project.json
    config_data = {
        "project_name": data.project_name,
        "project_path": actual_project_path,
        "project_description": data.project_description,
        "project_type": data.project_type,
    }
    config_service.write(config_data)
    
    return CreateProjectResponse(
        project_name=config_data["project_name"],
        project_path=config_data["project_path"],
        project_description=config_data["project_description"],
        project_type=config_data["project_type"],
        directory_name=dir_name,
    )


@router.patch("", response_model=ProjectSettings)
async def update_project_settings(
    request: fastapi.Request,
    update: ProjectSettingsUpdate,
):
    """Update project settings in project.json file."""
    settings = request.app.state.settings
    config_service = ProjectConfigService(settings.project_path)
    
    config_data = config_service.read()
    
    # Update only provided fields
    if update.project_name is not None:
        config_data["project_name"] = update.project_name
    if update.project_path is not None:
        config_data["project_path"] = update.project_path
    if update.project_description is not None:
        config_data["project_description"] = update.project_description
    if update.project_type is not None:
        config_data["project_type"] = update.project_type
    
    config_service.write(config_data)
    
    return ProjectSettings(
        project_name=config_data.get("project_name", ""),
        project_path=config_data.get("project_path", str(settings.project_path)),
        project_description=config_data.get("project_description", ""),
        project_type=config_data.get("project_type", "classification"),
    )


class DataManagerConfig(pydantic.BaseModel):
    """DataManager configuration model matching Python DataManager class."""
    test_size: float = 0.2
    n_splits: int = 5
    split_method: str = "shuffle"  # "shuffle" or "kfold"
    group_column: str | None = None
    stratified: bool = False
    random_state: int | None = None


class WriteDataFileRequest(pydantic.BaseModel):
    """Request model for writing data.py file."""
    base_data_manager: DataManagerConfig


class WriteDataFileResponse(pydantic.BaseModel):
    """Response model for data.py file creation."""
    success: bool
    file_path: str


@router.post("/data-file", response_model=WriteDataFileResponse)
async def write_data_file(
    request: fastapi.Request,
    data: WriteDataFileRequest,
):
    """Write the data.py file with BASE_DATA_MANAGER configuration.
    
    Creates a data.py file in the project directory with the DataManager
    configuration specified by the user.
    """
    settings = request.app.state.settings
    create_mode = os.environ.get("BRISK_UI_CREATE_MODE", "false") == "true"
    
    # In create mode, we need to find the actual project directory
    # by looking at what was created
    if create_mode:
        # Find the most recently created project directory
        subdirs = [d for d in settings.project_path.iterdir() if d.is_dir() and not d.name.startswith('.')]
        if subdirs:
            # Get most recent
            project_dir = max(subdirs, key=lambda d: d.stat().st_mtime)
        else:
            raise fastapi.HTTPException(
                status_code=400,
                detail="No project directory found. Create project first."
            )
    else:
        project_dir = settings.project_path
    
    data_file_path = project_dir / "data.py"
    
    # Build the DataManager instantiation string
    dm = data.base_data_manager
    params = []
    
    # Always include test_size
    params.append(f"    test_size = {dm.test_size}")
    
    # Include n_splits if not default
    if dm.n_splits != 5:
        params.append(f"    n_splits = {dm.n_splits}")
    
    # Include split_method if not default
    if dm.split_method != "shuffle":
        params.append(f'    split_method = "{dm.split_method}"')
    
    # Include group_column if set
    if dm.group_column:
        params.append(f'    group_column = "{dm.group_column}"')
    
    # Include stratified if True
    if dm.stratified:
        params.append(f"    stratified = {dm.stratified}")
    
    # Include random_state if set
    if dm.random_state is not None:
        params.append(f"    random_state = {dm.random_state}")
    
    params_str = ",\n".join(params)
    
    file_content = f'''# data.py
from brisk.data.data_manager import DataManager

BASE_DATA_MANAGER = DataManager(
{params_str}
)
'''
    
    try:
        with open(data_file_path, "w") as f:
            f.write(file_content)
        
        return WriteDataFileResponse(
            success=True,
            file_path=str(data_file_path)
        )
    except Exception as e:
        raise fastapi.HTTPException(
            status_code=500,
            detail=f"Failed to write data.py: {str(e)}"
        )


class AlgorithmWrapperConfig(pydantic.BaseModel):
    """Configuration for a single AlgorithmWrapper."""
    name: str  # Unique identifier
    display_name: str  # Human-readable name
    class_name: str  # Python class name (e.g., "Ridge")
    class_module: str  # Python module (e.g., "sklearn.linear_model")
    default_params: dict[str, str | int | float | bool | None] = {}
    search_space: dict[str, list[str | int | float | bool]] = {}  # Hyperparameter search space (arrays)
    use_defaults: bool = True


class WriteAlgorithmsFileRequest(pydantic.BaseModel):
    """Request model for writing algorithms.py file."""
    wrappers: list[AlgorithmWrapperConfig]


class WriteAlgorithmsFileResponse(pydantic.BaseModel):
    """Response model for algorithms.py file creation."""
    success: bool
    file_path: str


@router.post("/algorithms-file", response_model=WriteAlgorithmsFileResponse)
async def write_algorithms_file(
    request: fastapi.Request,
    data: WriteAlgorithmsFileRequest,
):
    """Write the algorithms.py file with ALGORITHM_CONFIG.
    
    Creates an algorithms.py file in the project directory with the
    AlgorithmWrapper instances specified by the user.
    """
    settings = request.app.state.settings
    create_mode = os.environ.get("BRISK_UI_CREATE_MODE", "false") == "true"
    
    # In create mode, we need to find the actual project directory
    if create_mode:
        subdirs = [d for d in settings.project_path.iterdir() if d.is_dir() and not d.name.startswith('.')]
        if subdirs:
            project_dir = max(subdirs, key=lambda d: d.stat().st_mtime)
        else:
            raise fastapi.HTTPException(
                status_code=400,
                detail="No project directory found. Create project first."
            )
    else:
        project_dir = settings.project_path
    
    algorithms_file_path = project_dir / "algorithms.py"
    
    # Collect unique imports needed
    imports_by_module: dict[str, set[str]] = {}
    for wrapper in data.wrappers:
        if wrapper.class_module not in imports_by_module:
            imports_by_module[wrapper.class_module] = set()
        imports_by_module[wrapper.class_module].add(wrapper.class_name)
    
    # Build import statements
    import_lines = ["from brisk.configuration.algorithm_wrapper import AlgorithmWrapper"]
    import_lines.append("import brisk")
    
    for module, classes in sorted(imports_by_module.items()):
        classes_str = ", ".join(sorted(classes))
        import_lines.append(f"from {module} import {classes_str}")
    
    imports_str = "\n".join(import_lines)
    
    # Build AlgorithmWrapper instantiations
    wrapper_strs = []
    for wrapper in data.wrappers:
        params_parts = [
            f'        name="{wrapper.name}"',
            f'        display_name="{wrapper.display_name}"',
            f'        algorithm_class={wrapper.class_name}',
        ]
        
        # Add default_params if there are any and not using defaults
        if wrapper.default_params and not wrapper.use_defaults:
            params_dict_str = _format_params_dict(wrapper.default_params)
            params_parts.append(f'        default_params={params_dict_str}')
        
        # Add search_space if there are any hyperparameter search values
        if wrapper.search_space:
            search_space_str = _format_search_space_dict(wrapper.search_space)
            if search_space_str != "{}":
                params_parts.append(f'        hyperparameters={search_space_str}')
        
        params_str = ",\n".join(params_parts)
        wrapper_strs.append(f"    AlgorithmWrapper(\n{params_str},\n    )")
    
    wrappers_str = ",\n".join(wrapper_strs)
    
    file_content = f'''# algorithms.py
{imports_str}

ALGORITHM_CONFIG = brisk.AlgorithmCollection(
{wrappers_str}
)
'''
    
    try:
        with open(algorithms_file_path, "w") as f:
            f.write(file_content)
        
        return WriteAlgorithmsFileResponse(
            success=True,
            file_path=str(algorithms_file_path)
        )
    except Exception as e:
        raise fastapi.HTTPException(
            status_code=500,
            detail=f"Failed to write algorithms.py: {str(e)}"
        )


def _format_params_dict(params: dict[str, str | int | float | bool | None]) -> str:
    """Format a params dict as a Python dict literal."""
    if not params:
        return "{}"
    
    parts = []
    for key, value in params.items():
        if isinstance(value, str):
            parts.append(f'"{key}": "{value}"')
        elif isinstance(value, bool):
            parts.append(f'"{key}": {value}')
        elif value is None:
            parts.append(f'"{key}": None')
        else:
            parts.append(f'"{key}": {value}')
    
    return "{" + ", ".join(parts) + "}"


def _format_search_space_dict(search_space: dict[str, list[str | int | float | bool]]) -> str:
    """Format a search space dict as a Python dict literal with lists.
    
    The search space values are arrays that define the hyperparameter search space
    for sklearn GridSearchCV or similar hyperparameter tuning.
    """
    if not search_space:
        return "{}"
    
    parts = []
    for key, values in search_space.items():
        if not values:  # Skip empty lists
            continue
        
        # Format the list values
        formatted_values = []
        for v in values:
            if isinstance(v, str):
                formatted_values.append(f'"{v}"')
            elif isinstance(v, bool):
                formatted_values.append(str(v))
            else:
                formatted_values.append(str(v))
        
        list_str = "[" + ", ".join(formatted_values) + "]"
        parts.append(f'"{key}": {list_str}')
    
    if not parts:
        return "{}"
    
    return "{" + ", ".join(parts) + "}"


def _get_project_dir(settings, create_mode: bool):
    """Get the project directory based on mode."""
    if create_mode:
        subdirs = [d for d in settings.project_path.iterdir() if d.is_dir() and not d.name.startswith('.')]
        if subdirs:
            return max(subdirs, key=lambda d: d.stat().st_mtime)
        else:
            raise fastapi.HTTPException(
                status_code=400,
                detail="No project directory found. Create project first."
            )
    return settings.project_path


class WriteMetricsFileRequest(pydantic.BaseModel):
    """Request model for writing metrics.py file."""
    problem_type: str  # "classification" or "regression"


class WriteMetricsFileResponse(pydantic.BaseModel):
    """Response model for metrics.py file creation."""
    success: bool
    file_path: str


@router.post("/metrics-file", response_model=WriteMetricsFileResponse)
async def write_metrics_file(
    request: fastapi.Request,
    data: WriteMetricsFileRequest,
):
    """Write the metrics.py file based on problem type.
    
    Creates a metrics.py file in the project directory with the default
    metrics for the specified problem type.
    """
    settings = request.app.state.settings
    create_mode = os.environ.get("BRISK_UI_CREATE_MODE", "false") == "true"
    
    project_dir = _get_project_dir(settings, create_mode)
    metrics_file_path = project_dir / "metrics.py"
    
    # Select metrics based on problem type
    if data.problem_type == "classification":
        metrics_line = "*brisk.CLASSIFICATION_METRICS"
    else:
        metrics_line = "*brisk.REGRESSION_METRICS"
    
    file_content = f'''# metrics.py
import brisk

METRIC_CONFIG = brisk.MetricManager(
    {metrics_line}
)
'''
    
    try:
        with open(metrics_file_path, "w") as f:
            f.write(file_content)
        
        return WriteMetricsFileResponse(
            success=True,
            file_path=str(metrics_file_path)
        )
    except Exception as e:
        raise fastapi.HTTPException(
            status_code=500,
            detail=f"Failed to write metrics.py: {str(e)}"
        )


class WriteEvaluatorsFileResponse(pydantic.BaseModel):
    """Response model for evaluators.py file creation."""
    success: bool
    file_path: str


@router.post("/evaluators-file", response_model=WriteEvaluatorsFileResponse)
async def write_evaluators_file(request: fastapi.Request):
    """Write the evaluators.py file (template).
    
    Creates a boilerplate evaluators.py file in the project directory
    for custom evaluator registration.
    """
    settings = request.app.state.settings
    create_mode = os.environ.get("BRISK_UI_CREATE_MODE", "false") == "true"
    
    project_dir = _get_project_dir(settings, create_mode)
    evaluators_file_path = project_dir / "evaluators.py"
    
    file_content = '''# evaluators.py
# Define custom evaluation methods here to integrate with Brisk's builtin tools
from brisk.evaluation.evaluators.registry import EvaluatorRegistry
from brisk import PlotEvaluator, MeasureEvaluator

def register_custom_evaluators(registry: EvaluatorRegistry, plot_settings) -> None:
    # registry.register(
    # Initalize an evaluator instance here to register
    # )
    pass
'''
    
    try:
        with open(evaluators_file_path, "w") as f:
            f.write(file_content)
        
        return WriteEvaluatorsFileResponse(
            success=True,
            file_path=str(evaluators_file_path)
        )
    except Exception as e:
        raise fastapi.HTTPException(
            status_code=500,
            detail=f"Failed to write evaluators.py: {str(e)}"
        )


# --- Workflow file ---

class WorkflowStepConfig(pydantic.BaseModel):
    """A single workflow step (evaluator call)."""
    evaluator_id: str
    method_name: str
    args: dict = {}


class WriteWorkflowFileRequest(pydantic.BaseModel):
    """Request model for writing workflows/<problem_type>.py file."""
    problem_type: str  # "classification" or "regression"
    steps: list[WorkflowStepConfig]


class WriteWorkflowFileResponse(pydantic.BaseModel):
    """Response model for workflow file creation."""
    success: bool
    file_path: str


# Methods that take X, y as np.ndarray (emit X.values, y.values)
_WORKFLOW_ARRAY_DATA_METHODS = frozenset({
    "confusion_matrix",
    "plot_confusion_heatmap",
    "plot_roc_curve",
    "plot_precision_recall_curve",
})


def _format_workflow_step_call(step: WorkflowStepConfig, problem_type: str) -> str:
    """Format a single workflow step as a Python method call.
    Uses self.model for model arg; X_train, X_test, y_train, y_test from workflow params.
    """
    method = step.method_name
    args = step.args or {}

    def get_data_var(key: str, default: str = "X_test") -> str:
        v = args.get(key)
        if v is None or v == "":
            return default
        return str(v)

    if method == "evaluate_model":
        X = get_data_var("X", "X_test")
        y = get_data_var("y", "y_test")
        metrics_raw = args.get("metrics") or "MAE"
        if isinstance(metrics_raw, list):
            metrics_str = ", ".join(f'"{m}"' for m in metrics_raw)
        else:
            parts = [s.strip() for s in str(metrics_raw).split(",") if s.strip()]
            metrics_str = ", ".join(f'"{p}"' for p in parts) if parts else '"MAE"'
        filename = args.get("filename") or "evaluate_model"
        return f'        self.evaluate_model(self.model, {X}, {y}, [{metrics_str}], "{filename}")'
    if method == "evaluate_model_cv":
        X = get_data_var("X", "X_train")
        y = get_data_var("y", "y_train")
        metrics_raw = args.get("metrics") or "MAE"
        if isinstance(metrics_raw, list):
            metrics_str = ", ".join(f'"{m}"' for m in metrics_raw)
        else:
            parts = [s.strip() for s in str(metrics_raw).split(",") if s.strip()]
            metrics_str = ", ".join(f'"{p}"' for p in parts) if parts else '"MAE"'
        filename = args.get("filename") or "evaluate_model_cv"
        cv = args.get("cv")
        cv_val = int(cv) if cv is not None and str(cv).strip() != "" else 5
        return f'        self.evaluate_model_cv(self.model, {X}, {y}, [{metrics_str}], "{filename}", cv={cv_val})'
    if method == "plot_pred_vs_obs":
        X = get_data_var("X", "X_test")
        y = get_data_var("y", "y_test")
        filename = args.get("filename") or "pred_vs_obs"
        return f'        self.plot_pred_vs_obs(self.model, {X}, {y}, "{filename}")'
    if method == "plot_learning_curve":
        X_train = get_data_var("X", "X_train")
        y_train = get_data_var("y", "y_train")
        filename = args.get("filename") or "learning_curve"
        cv = args.get("cv")
        cv_val = int(cv) if cv is not None and str(cv).strip() != "" else 5
        num_repeats = args.get("num_repeats")
        num_repeats_val = int(num_repeats) if num_repeats is not None and str(num_repeats).strip() != "" else 1
        n_jobs = args.get("n_jobs")
        n_jobs_val = int(n_jobs) if n_jobs is not None and str(n_jobs).strip() != "" else -1
        metric = args.get("metric") or "neg_mean_absolute_error"
        return f'        self.plot_learning_curve(self.model, {X_train}, {y_train}, filename="{filename}", cv={cv_val}, num_repeats={num_repeats_val}, n_jobs={n_jobs_val}, metric="{metric}")'
    if method == "plot_feature_importance":
        X = get_data_var("X", "X_train")
        y = get_data_var("y", "y_train")
        threshold = args.get("threshold")
        thresh_val = int(threshold) if threshold is not None and str(threshold).strip() != "" else 10
        filename = args.get("filename") or "feature_importance"
        metric = args.get("metric") or "neg_mean_absolute_error"
        num_rep = args.get("num_rep")
        num_rep_val = int(num_rep) if num_rep is not None and str(num_rep).strip() != "" else 5
        return f'        self.plot_feature_importance(self.model, {X}, {y}, {thresh_val}, feature_names, "{filename}", "{metric}", {num_rep_val})'
    if method == "plot_residuals":
        X = get_data_var("X", "X_test")
        y = get_data_var("y", "y_test")
        filename = args.get("filename") or "residuals"
        add_fit = args.get("add_fit_line")
        add_fit_val = "True" if add_fit else "False"
        return f'        self.plot_residuals(self.model, {X}, {y}, "{filename}", add_fit_line={add_fit_val})'
    if method == "hyperparameter_tuning":
        X_train = get_data_var("X", "X_train")
        y_train = get_data_var("y", "y_train")
        method_type = args.get("method") or "grid"
        scorer = args.get("scorer") or "neg_mean_absolute_error"
        kf = args.get("kf")
        kf_val = int(kf) if kf is not None and str(kf).strip() != "" else 5
        num_rep = args.get("num_rep")
        num_rep_val = int(num_rep) if num_rep is not None and str(num_rep).strip() != "" else 3
        n_jobs = args.get("n_jobs")
        n_jobs_val = int(n_jobs) if n_jobs is not None and str(n_jobs).strip() != "" else -1
        plot_res = args.get("plot_results")
        plot_res_val = "True" if plot_res else "False"
        return f'        self.hyperparameter_tuning(self.model, "{method_type}", {X_train}, {y_train}, "{scorer}", {kf_val}, {num_rep_val}, {n_jobs_val}, plot_results={plot_res_val})'
    if method == "confusion_matrix":
        X = get_data_var("X", "X_test")
        y = get_data_var("y", "y_test")
        filename = args.get("filename") or "confusion_matrix"
        return f'        self.confusion_matrix(self.model, {X}.values, {y}.values, "{filename}")'
    if method == "plot_confusion_heatmap":
        X = get_data_var("X", "X_test")
        y = get_data_var("y", "y_test")
        filename = args.get("filename") or "confusion_heatmap"
        return f'        self.plot_confusion_heatmap(self.model, {X}.values, {y}.values, "{filename}")'
    if method == "plot_roc_curve":
        X = get_data_var("X", "X_test")
        y = get_data_var("y", "y_test")
        filename = args.get("filename") or "roc_curve"
        pos_label = args.get("pos_label")
        pos_val = int(pos_label) if pos_label is not None and str(pos_label).strip() != "" else 1
        return f'        self.plot_roc_curve(self.model, {X}.values, {y}.values, "{filename}", pos_label={pos_val})'
    if method == "plot_precision_recall_curve":
        X = get_data_var("X", "X_test")
        y = get_data_var("y", "y_test")
        filename = args.get("filename") or "precision_recall_curve"
        pos_label = args.get("pos_label")
        pos_val = int(pos_label) if pos_label is not None and str(pos_label).strip() != "" else 1
        return f'        self.plot_precision_recall_curve(self.model, {X}.values, {y}.values, "{filename}", pos_label={pos_val})'
    if method == "plot_shapley_values":
        X = get_data_var("X", "X_test")
        y = get_data_var("y", "y_test")
        filename = args.get("filename") or "shapley_values"
        plot_type = args.get("plot_type") or "bar"
        return f'        self.plot_shapley_values(self.model, {X}, {y}, filename="{filename}", plot_type="{plot_type}")'
    if method == "save_model":
        filename = args.get("filename") or "model"
        return f'        self.save_model(self.model, "{filename}")'
    return ""


@router.post("/workflow-file", response_model=WriteWorkflowFileResponse)
async def write_workflow_file(
    request: fastapi.Request,
    data: WriteWorkflowFileRequest,
):
    """Write the workflow file to workflows/<problem_type>.py.
    
    Creates a workflow file with class Regression or Classification
    and def workflow(self, X_train, X_test, y_train, y_test, output_dir, feature_names).
    """
    settings = request.app.state.settings
    create_mode = os.environ.get("BRISK_UI_CREATE_MODE", "false") == "true"
    project_dir = _get_project_dir(settings, create_mode)
    workflows_dir = project_dir / "workflows"
    workflows_dir.mkdir(parents=True, exist_ok=True)
    problem_type = data.problem_type
    if problem_type not in ("classification", "regression"):
        problem_type = "regression"
    class_name = problem_type.capitalize()
    workflow_file_path = workflows_dir / f"{problem_type}.py"

    step_lines = []
    for step in data.steps:
        line = _format_workflow_step_call(step, problem_type)
        if line:
            step_lines.append(line)

    body = "\n".join(step_lines) if step_lines else "        pass"

    file_content = f'''# workflow.py
# Define the workflow for training and evaluating models

from brisk.training.workflow import Workflow


class {class_name}(Workflow):
    def workflow(self, X_train, X_test, y_train, y_test, output_dir, feature_names):
{body}
'''
    try:
        with open(workflow_file_path, "w") as f:
            f.write(file_content)
        return WriteWorkflowFileResponse(
            success=True,
            file_path=str(workflow_file_path)
        )
    except Exception as e:
        raise fastapi.HTTPException(
            status_code=500,
            detail=f"Failed to write workflow file: {str(e)}"
        )


class ExperimentGroupDataConfig(pydantic.BaseModel):
    """DataManager config for an experiment group (data_config parameter)."""
    test_size: float | None = None
    n_splits: int | None = None
    split_method: str | None = None
    group_column: str | None = None
    stratified: bool | None = None
    random_state: int | None = None
    preprocessors: list[dict] | None = None  # list of {type, config} for each preprocessor


class ExperimentGroupConfig(pydantic.BaseModel):
    """Configuration for a single experiment group."""
    name: str
    description: str = ""
    dataset_file_name: str
    dataset_table_name: str | None = None
    algorithms: list[str]
    use_default_data_manager: bool = True
    data_config: ExperimentGroupDataConfig | None = None


class PlotSettingsPayload(pydantic.BaseModel):
    """PlotSettings payload; only non-default values are written."""
    file_format: str | None = None
    transparent: bool | None = None
    width: int | None = None
    height: int | None = None
    dpi: int | None = None
    primary_color: str | None = None
    secondary_color: str | None = None
    accent_color: str | None = None


class CategoricalFeaturesEntry(pydantic.BaseModel):
    """Entry for categorical features mapping."""
    dataset_file_name: str  # e.g., "data.csv" or "data.sqlite"
    table_name: str | None = None  # for SQLite, the table name
    features: list[str]  # list of categorical feature names


class WriteSettingsFileRequest(pydantic.BaseModel):
    """Request model for writing settings.py file."""
    problem_type: str  # "classification" or "regression"
    default_algorithms: list[str]  # list of all algorithm names
    experiment_groups: list[ExperimentGroupConfig]
    categorical_features: list[CategoricalFeaturesEntry] | None = None  # optional categorical features mapping
    plot_settings: PlotSettingsPayload | None = None


class WriteSettingsFileResponse(pydantic.BaseModel):
    """Response model for settings.py file creation."""
    success: bool
    file_path: str


def _format_preprocessor_instance(preproc_type: str, config: dict, problem_type: str) -> str:
    """Format a single preprocessor as Python instantiation code.
    Maps frontend preprocessor types/config to Brisk class constructors.
    Brisk pipeline order: Missing Data -> Categorical Encoding -> Scaling -> Feature Selection.
    """
    if preproc_type == "missing-data":
        strategy = config.get("strategy", "mean")
        if strategy == "drop":
            return "MissingDataPreprocessor(strategy=\"drop_rows\")"
        # strategy is impute with impute_method
        impute_method = strategy if strategy != "most_frequent" else "mode"
        constant_value = config.get("fillValue")
        if impute_method == "constant" and constant_value is not None:
            try:
                cv = float(constant_value)
                return f"MissingDataPreprocessor(strategy=\"impute\", impute_method=\"constant\", constant_value={cv})"
            except (TypeError, ValueError):
                return f"MissingDataPreprocessor(strategy=\"impute\", impute_method=\"constant\", constant_value=\"{constant_value}\")"
        return f"MissingDataPreprocessor(strategy=\"impute\", impute_method=\"{impute_method}\")"
    if preproc_type == "scaling":
        method = config.get("method", "standard")
        return f"ScalingPreprocessor(method=\"{method}\")"
    if preproc_type == "encoding":
        method = config.get("method", "label")
        if method == "target":
            method = "label"
        return f"CategoricalEncodingPreprocessor(method=\"{method}\")"
    if preproc_type == "feature-selection":
        method = config.get("method", "selectkbest")
        # Map frontend method to Brisk: variance|univariate -> selectkbest, recursive -> rfecv, lasso -> sequential
        brisk_method = {"variance": "selectkbest", "univariate": "selectkbest", "recursive": "rfecv", "lasso": "sequential"}.get(method, "selectkbest")
        n_features = config.get("nFeatures", 5)
        if n_features == "auto":
            n_features = 5
        try:
            n_features = int(n_features)
        except (TypeError, ValueError):
            n_features = 5
        return f"FeatureSelectionPreprocessor(method=\"{brisk_method}\", n_features_to_select={n_features}, feature_selection_cv=3, problem_type=\"{problem_type}\")"
    return ""


def _format_preprocessors_list(preprocessors: list[dict], problem_type: str) -> str | None:
    """Format preprocessors list for data_config. Returns Python code for the list or None."""
    if not preprocessors:
        return None
    # Brisk pipeline order: missing-data -> encoding -> scaling -> feature-selection
    order = ("missing-data", "encoding", "scaling", "feature-selection")
    ordered = []
    for key in order:
        for p in preprocessors:
            if p.get("type") == key:
                code = _format_preprocessor_instance(key, p.get("config") or {}, problem_type)
                if code:
                    ordered.append(f"            {code},")
                break
    if not ordered:
        return None
    return "[\n" + "\n".join(ordered) + "\n        ]"


@router.post("/settings-file", response_model=WriteSettingsFileResponse)
async def write_settings_file(
    request: fastapi.Request,
    data: WriteSettingsFileRequest,
):
    """Write the settings.py file with Configuration.
    
    Creates a settings.py file in the project directory with the
    Configuration and experiment groups.
    """
    settings = request.app.state.settings
    create_mode = os.environ.get("BRISK_UI_CREATE_MODE", "false") == "true"
    
    project_dir = _get_project_dir(settings, create_mode)
    settings_file_path = project_dir / "settings.py"
    
    # PlotSettings: only include non-default kwargs (Brisk PlotSettings defaults)
    _plot_defaults = {
        "file_format": "png",
        "transparent": False,
        "width": 10,
        "height": 8,
        "dpi": 300,
        "primary_color": "#1175D5",
        "secondary_color": "#00A878",
        "accent_color": "#DE6B48",
    }
    plot_kwargs: dict[str, str | int | bool] = {}
    if data.plot_settings:
        ps = data.plot_settings
        if ps.file_format is not None and ps.file_format != _plot_defaults["file_format"]:
            plot_kwargs["file_format"] = ps.file_format
        if ps.transparent is not None and ps.transparent != _plot_defaults["transparent"]:
            plot_kwargs["transparent"] = ps.transparent
        if ps.width is not None and ps.width != _plot_defaults["width"]:
            plot_kwargs["width"] = ps.width
        if ps.height is not None and ps.height != _plot_defaults["height"]:
            plot_kwargs["height"] = ps.height
        if ps.dpi is not None and ps.dpi != _plot_defaults["dpi"]:
            plot_kwargs["dpi"] = ps.dpi
        if ps.primary_color is not None and ps.primary_color != _plot_defaults["primary_color"]:
            plot_kwargs["primary_color"] = ps.primary_color
        if ps.secondary_color is not None and ps.secondary_color != _plot_defaults["secondary_color"]:
            plot_kwargs["secondary_color"] = ps.secondary_color
        if ps.accent_color is not None and ps.accent_color != _plot_defaults["accent_color"]:
            plot_kwargs["accent_color"] = ps.accent_color

    # Format algorithm list
    algorithms_str = ", ".join(f'"{a}"' for a in data.default_algorithms)
    
    # Check if any group needs preprocessor imports
    needs_preprocessors_import = any(
        group.data_config and group.data_config.preprocessors
        for group in data.experiment_groups
    )
    
    # Build experiment group calls
    group_strs = []
    for group in data.experiment_groups:
        parts = [f'        name="{group.name}"']
        
        if group.description:
            parts.append(f'        description="{group.description}"')
        
        # Format datasets - tuple if table name exists
        if group.dataset_table_name:
            parts.append(f'        datasets=[("{group.dataset_file_name}", "{group.dataset_table_name}")]')
        else:
            parts.append(f'        datasets=["{group.dataset_file_name}"]')
        
        # Algorithms
        alg_str = ", ".join(f'"{a}"' for a in group.algorithms)
        parts.append(f'        algorithms=[{alg_str}]')
        
        # data_config - include when not using default data manager params, or when preprocessors are present
        dc = group.data_config
        has_dc_params = dc and (
            dc.test_size is not None or dc.n_splits is not None or dc.split_method is not None
            or dc.group_column is not None or dc.stratified is not None or dc.random_state is not None
        )
        has_preprocessors = dc and dc.preprocessors
        if has_preprocessors or (not group.use_default_data_manager and has_dc_params):
            dc_parts = []
            if dc:
                if dc.test_size is not None:
                    dc_parts.append(f'"test_size": {dc.test_size}')
                if dc.n_splits is not None:
                    dc_parts.append(f'"n_splits": {dc.n_splits}')
                if dc.split_method is not None:
                    dc_parts.append(f'"split_method": "{dc.split_method}"')
                if dc.group_column is not None:
                    dc_parts.append(f'"group_column": "{dc.group_column}"')
                if dc.stratified is not None:
                    dc_parts.append(f'"stratified": {dc.stratified}')
                if dc.random_state is not None:
                    dc_parts.append(f'"random_state": {dc.random_state}')
                preprocessors_code = _format_preprocessors_list(dc.preprocessors or [], data.problem_type)
                if preprocessors_code:
                    dc_parts.append(f'"preprocessors": {preprocessors_code}')
            if dc_parts:
                dc_str = ", ".join(dc_parts)
                parts.append(f'        data_config={{{dc_str}}}')
        
        params_str = ",\n".join(parts)
        group_strs.append(f"    config.add_experiment_group(\n{params_str},\n    )")
    
    groups_code = "\n\n".join(group_strs)
    
    preprocessors_import = ""
    if needs_preprocessors_import:
        preprocessors_import = "\nfrom brisk.data.preprocessing import (\n    MissingDataPreprocessor,\n    ScalingPreprocessor,\n    CategoricalEncodingPreprocessor,\n    FeatureSelectionPreprocessor,\n)\n"

    plot_settings_import = ""
    plot_settings_var = ""
    config_plot_arg = ""
    if plot_kwargs:
        plot_settings_import = "\nfrom brisk.theme.plot_settings import PlotSettings\n"
        kwargs_str = ", ".join(
            f'{k}="{v}"' if isinstance(v, str) else f"{k}={v}"
            for k, v in plot_kwargs.items()
        )
        plot_settings_var = f"\n    plot_settings = PlotSettings({kwargs_str})\n"
        config_plot_arg = ",\n        plot_settings=plot_settings"

    # Build categorical_features dict if provided
    categorical_features_arg = ""
    if data.categorical_features:
        cat_entries = []
        for entry in data.categorical_features:
            if not entry.features:  # skip if no features
                continue
            features_str = ", ".join(f'"{f}"' for f in entry.features)
            if entry.table_name:
                # SQLite: key is tuple (filename, table_name)
                cat_entries.append(f'            ("{entry.dataset_file_name}", "{entry.table_name}"): [{features_str}]')
            else:
                # CSV/XLSX: key is just filename
                cat_entries.append(f'            "{entry.dataset_file_name}": [{features_str}]')
        if cat_entries:
            categorical_features_arg = ",\n        categorical_features={\n" + ",\n".join(cat_entries) + "\n        }"

    file_content = f'''# settings.py
from brisk.configuration.configuration import Configuration
from brisk.configuration.configuration_manager import ConfigurationManager
{preprocessors_import}{plot_settings_import}

def create_configuration() -> ConfigurationManager:
{plot_settings_var}    config = Configuration(
        default_workflow="{data.problem_type}",
        default_algorithms=[{algorithms_str}]{categorical_features_arg}{config_plot_arg}
    )

{groups_code}

    return config.build()
'''
    
    try:
        with open(settings_file_path, "w") as f:
            f.write(file_content)
        
        return WriteSettingsFileResponse(
            success=True,
            file_path=str(settings_file_path)
        )
    except Exception as e:
        raise fastapi.HTTPException(
            status_code=500,
            detail=f"Failed to write settings.py: {str(e)}"
        )


class DeleteResponse(pydantic.BaseModel):
    """Response model for project deletion."""
    success: bool
    message: str
    deleted_path: str


@router.delete("", response_model=DeleteResponse)
async def delete_project(request: fastapi.Request):
    """Delete the entire project directory.
    
    WARNING: This is a destructive operation that cannot be undone.
    It deletes the entire project directory, not just the .brisk folder.
    """
    settings = request.app.state.settings
    project_path = settings.project_path
    
    if not project_path.exists():
        raise fastapi.HTTPException(
            status_code=404,
            detail=f"Project directory not found: {project_path}"
        )
    
    # Safety check: ensure we're deleting a brisk project
    brisk_dir = project_path / ".brisk"
    if not brisk_dir.exists():
        raise fastapi.HTTPException(
            status_code=400,
            detail="Not a valid brisk project (missing .brisk directory)"
        )
    
    try:
        deleted_path = str(project_path)
        shutil.rmtree(project_path)
        return DeleteResponse(
            success=True,
            message="Project deleted successfully",
            deleted_path=deleted_path
        )
    except PermissionError:
        raise fastapi.HTTPException(
            status_code=403,
            detail="Permission denied: cannot delete project directory"
        )
    except Exception as e:
        raise fastapi.HTTPException(
            status_code=500,
            detail=f"Failed to delete project: {str(e)}"
        )


class ProjectStats(pydantic.BaseModel):
    """Project statistics model."""
    groups: int = 0
    experiments: int = 0
    datasets: int = 0
    algorithms: int = 0
    metrics: int = 0


def _count_experiment_groups_and_experiments(settings_content: str) -> tuple[int, int]:
    """Parse settings.py to count experiment groups and calculate total experiments.
    
    An experiment = number of datasets × number of algorithms in each group.
    """
    # Count add_experiment_group calls
    group_pattern = r'config\.add_experiment_group\s*\('
    groups = len(re.findall(group_pattern, settings_content))
    
    # Find each add_experiment_group call and extract datasets/algorithms
    total_experiments = 0
    
    # Pattern to match add_experiment_group with its arguments
    # This is a simplified parser - looks for datasets=[...] and algorithms=[...]
    group_calls = re.finditer(
        r'config\.add_experiment_group\s*\((.*?)\)\s*(?=config\.|return|\Z)',
        settings_content,
        re.DOTALL
    )
    
    for match in group_calls:
        call_content = match.group(1)
        
        # Count datasets - look for datasets=[...] or datasets=["..."]
        datasets_match = re.search(r'datasets\s*=\s*\[(.*?)\]', call_content, re.DOTALL)
        if datasets_match:
            datasets_str = datasets_match.group(1)
            # Count items (strings or tuples)
            # Simple approach: count quoted strings or tuple patterns
            dataset_items = re.findall(r'(?:"[^"]*"|\'[^\']*\'|\([^)]+\))', datasets_str)
            num_datasets = len(dataset_items) if dataset_items else 0
        else:
            num_datasets = 0
        
        # Count algorithms - look for algorithms=[...]
        algorithms_match = re.search(r'algorithms\s*=\s*\[(.*?)\]', call_content, re.DOTALL)
        if algorithms_match:
            algorithms_str = algorithms_match.group(1)
            # Count quoted strings
            algorithm_items = re.findall(r'(?:"[^"]*"|\'[^\']*\')', algorithms_str)
            num_algorithms = len(algorithm_items) if algorithm_items else 0
        else:
            num_algorithms = 0
        
        total_experiments += num_datasets * num_algorithms
    
    return groups, total_experiments


def _count_datasets(project_path) -> int:
    """Count dataset files (.csv, .xlsx) and sqlite tables in datasets/ directory."""
    import sqlite3
    
    datasets_dir = project_path / "datasets"
    count = 0
    
    if not datasets_dir.exists():
        return 0
    
    for item in datasets_dir.iterdir():
        if item.is_file():
            suffix = item.suffix.lower()
            if suffix in ('.csv', '.xlsx', '.xls'):
                count += 1
            elif suffix in ('.sqlite', '.db', '.sqlite3'):
                # Count tables in sqlite database
                try:
                    conn = sqlite3.connect(str(item))
                    cursor = conn.execute(
                        "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
                    )
                    table_count = cursor.fetchone()[0]
                    count += table_count
                    conn.close()
                except Exception:
                    pass  # Skip if can't read database
    
    return count


def _count_algorithm_wrappers(algorithms_content: str) -> int:
    """Count AlgorithmWrapper instances in algorithms.py."""
    # Count AlgorithmWrapper( occurrences
    pattern = r'AlgorithmWrapper\s*\('
    return len(re.findall(pattern, algorithms_content))


def _count_metrics(metrics_content: str) -> int:
    """Count metrics in metrics.py.
    
    Handles both explicit metric definitions and brisk.REGRESSION_METRICS/CLASSIFICATION_METRICS.
    """
    count = 0
    
    # Check for brisk built-in metric collections
    if '*brisk.REGRESSION_METRICS' in metrics_content or '*brisk.regression_metrics' in metrics_content.lower():
        # Brisk has 6 default regression metrics
        count += 6
    if '*brisk.CLASSIFICATION_METRICS' in metrics_content or '*brisk.classification_metrics' in metrics_content.lower():
        # Brisk has 5 default classification metrics
        count += 5
    
    # Count explicit Metric(...) or MetricWrapper(...) definitions
    explicit_pattern = r'(?:Metric|MetricWrapper)\s*\('
    count += len(re.findall(explicit_pattern, metrics_content))
    
    return count


@router.get("/stats", response_model=ProjectStats)
async def get_project_stats(request: fastapi.Request):
    """Get project statistics by parsing project files.
    
    Returns counts for:
    - groups: Number of experiment groups in settings.py
    - experiments: Total experiments (datasets × algorithms per group)
    - datasets: Number of dataset files (.csv, .xlsx) + sqlite tables
    - algorithms: Number of AlgorithmWrapper instances in algorithms.py
    - metrics: Number of metrics in metrics.py
    """
    settings = request.app.state.settings
    project_path = settings.project_path
    
    stats = ProjectStats()
    
    # Parse settings.py for groups and experiments
    settings_file = project_path / "settings.py"
    if settings_file.exists():
        try:
            content = settings_file.read_text()
            stats.groups, stats.experiments = _count_experiment_groups_and_experiments(content)
        except Exception:
            pass
    
    # Count datasets
    stats.datasets = _count_datasets(project_path)
    
    # Parse algorithms.py
    algorithms_file = project_path / "algorithms.py"
    if algorithms_file.exists():
        try:
            content = algorithms_file.read_text()
            stats.algorithms = _count_algorithm_wrappers(content)
        except Exception:
            pass
    
    # Parse metrics.py
    metrics_file = project_path / "metrics.py"
    if metrics_file.exists():
        try:
            content = metrics_file.read_text()
            stats.metrics = _count_metrics(content)
        except Exception:
            pass
    
    return stats


# ============================================================================
# File Preview Endpoints
# ============================================================================

class ProjectFileInfo(pydantic.BaseModel):
    """Information about a project file."""
    id: str
    name: str
    path: str
    exists: bool
    size: int = 0


class ProjectFilesResponse(pydantic.BaseModel):
    """Response containing list of project files."""
    files: list[ProjectFileInfo]
    project_type: str


class FileContentResponse(pydantic.BaseModel):
    """Response containing file content."""
    name: str
    content: str
    exists: bool


# The standard project files to show (workflow file name depends on project_type)
PROJECT_FILES = [
    ("settings.py", "settings.py"),
    ("algorithms.py", "algorithms.py"),
    ("metrics.py", "metrics.py"),
    ("data.py", "data.py"),
    ("evaluators.py", "evaluators.py"),
]


@router.get("/files", response_model=ProjectFilesResponse)
async def get_project_files(request: fastapi.Request):
    """Get list of project configuration files with their status.
    
    Returns the standard project files plus the workflow file
    (regression.py or classification.py based on project type).
    """
    import pathlib
    
    settings = request.app.state.settings
    project_path = settings.project_path
    
    # Get project type from project.json
    config_service = ProjectConfigService(project_path)
    config_data = config_service.read()
    project_type = config_data.get("project_type", "classification")
    
    files: list[ProjectFileInfo] = []
    
    # Add standard files
    for file_id, filename in PROJECT_FILES:
        file_path = project_path / filename
        exists = file_path.exists()
        size = file_path.stat().st_size if exists else 0
        files.append(ProjectFileInfo(
            id=file_id,
            name=filename,
            path=str(file_path),
            exists=exists,
            size=size,
        ))
    
    # Add workflow file based on project type
    workflow_filename = f"{project_type}.py"
    workflow_path = project_path / "workflows" / workflow_filename
    exists = workflow_path.exists()
    size = workflow_path.stat().st_size if exists else 0
    files.append(ProjectFileInfo(
        id="workflow",
        name=workflow_filename,
        path=str(workflow_path),
        exists=exists,
        size=size,
    ))
    
    return ProjectFilesResponse(files=files, project_type=project_type)


@router.get("/files/{file_id}/content", response_model=FileContentResponse)
async def get_file_content(file_id: str, request: fastapi.Request):
    """Get the content of a specific project file.
    
    file_id can be: settings.py, algorithms.py, metrics.py, data.py, evaluators.py, or workflow
    """
    import pathlib
    
    settings = request.app.state.settings
    project_path = settings.project_path
    
    # Get project type for workflow file
    config_service = ProjectConfigService(project_path)
    config_data = config_service.read()
    project_type = config_data.get("project_type", "classification")
    
    # Determine the file path
    if file_id == "workflow":
        workflow_filename = f"{project_type}.py"
        file_path = project_path / "workflows" / workflow_filename
        name = workflow_filename
    else:
        # file_id is the filename itself (e.g., "settings.py")
        file_path = project_path / file_id
        name = file_id
    
    if not file_path.exists():
        return FileContentResponse(
            name=name,
            content="",
            exists=False,
        )
    
    try:
        content = file_path.read_text(encoding="utf-8")
        return FileContentResponse(
            name=name,
            content=content,
            exists=True,
        )
    except Exception as e:
        raise fastapi.HTTPException(
            status_code=500,
            detail=f"Failed to read file: {str(e)}"
        )


@router.post("/files/{file_id}/download")
async def download_file(file_id: str, request: fastapi.Request):
    """Download a project file.
    
    Returns the file as an attachment for download.
    """
    from fastapi.responses import FileResponse
    import pathlib
    
    settings = request.app.state.settings
    project_path = settings.project_path
    
    # Get project type for workflow file
    config_service = ProjectConfigService(project_path)
    config_data = config_service.read()
    project_type = config_data.get("project_type", "classification")
    
    # Determine the file path
    if file_id == "workflow":
        workflow_filename = f"{project_type}.py"
        file_path = project_path / "workflows" / workflow_filename
    else:
        file_path = project_path / file_id
    
    if not file_path.exists():
        raise fastapi.HTTPException(
            status_code=404,
            detail=f"File not found: {file_id}"
        )
    
    return FileResponse(
        path=str(file_path),
        filename=file_path.name,
        media_type="text/x-python",
    )


# ============================================================================
# Experiments Data Endpoints
# ============================================================================

class DatasetInfo(pydantic.BaseModel):
    """Information about a dataset file."""
    name: str  # filename without extension (used as identifier)
    filename: str  # full filename
    file_type: str  # csv, xlsx, sqlite, etc.


class AlgorithmInfo(pydantic.BaseModel):
    """Information about a configured algorithm."""
    name: str  # internal name
    display_name: str
    class_name: str  # Python class name (e.g., "Ridge")
    class_module: str  # Python module (e.g., "sklearn.linear_model")
    default_params: dict = {}  # Default parameters
    use_defaults: bool = True  # Whether using catalog defaults


class ExperimentGroupInfo(pydantic.BaseModel):
    """Information about an experiment group from settings.py."""
    name: str
    description: str
    datasets: list[str]
    algorithms: list[str]


class ExperimentsDataResponse(pydantic.BaseModel):
    """Response containing all data needed for experiments page."""
    datasets: list[DatasetInfo]
    algorithms: list[AlgorithmInfo]
    experiment_groups: list[ExperimentGroupInfo]


def _parse_algorithms_from_file(content: str) -> list[AlgorithmInfo]:
    """Parse algorithm info from algorithms.py content."""
    algorithms = []
    
    # Parse import statements to build a module map
    # Pattern: from sklearn.xxx import ClassName1, ClassName2
    import_pattern = r'from\s+([\w.]+)\s+import\s+([^#\n]+)'
    import_matches = re.findall(import_pattern, content)
    
    class_to_module = {}
    for module, classes_str in import_matches:
        # Split by comma and strip whitespace
        classes = [c.strip() for c in classes_str.split(',')]
        for class_name in classes:
            if class_name:
                class_to_module[class_name] = module
    
    # Match individual AlgorithmWrapper instantiations
    # Pattern captures the entire AlgorithmWrapper(...) block
    wrapper_pattern = r'AlgorithmWrapper\s*\(([\s\S]*?)\)'
    wrapper_matches = re.finditer(wrapper_pattern, content)
    
    for match in wrapper_matches:
        wrapper_content = match.group(1)
        
        # Extract name
        name_match = re.search(r'name\s*=\s*["\']([^"\']+)["\']', wrapper_content)
        name = name_match.group(1) if name_match else ""
        
        # Extract display_name
        display_name_match = re.search(r'display_name\s*=\s*["\']([^"\']+)["\']', wrapper_content)
        display_name = display_name_match.group(1) if display_name_match else name
        
        # Extract algorithm_class (the class reference, not quoted)
        class_match = re.search(r'algorithm_class\s*=\s*(\w+)', wrapper_content)
        class_name = class_match.group(1) if class_match else ""
        
        # Get module from our import map
        class_module = class_to_module.get(class_name, "sklearn")
        
        # Extract default_params if present
        default_params = {}
        params_match = re.search(r'default_params\s*=\s*\{([^}]*)\}', wrapper_content)
        if params_match:
            params_str = params_match.group(1)
            # Parse simple key-value pairs
            param_pattern = r'["\'](\w+)["\']\s*:\s*([^,}]+)'
            param_matches = re.findall(param_pattern, params_str)
            for param_name, param_value in param_matches:
                # Try to parse the value
                param_value = param_value.strip()
                if param_value.startswith('"') or param_value.startswith("'"):
                    # String value
                    default_params[param_name] = param_value.strip('"\'')
                elif param_value.lower() == 'true':
                    default_params[param_name] = True
                elif param_value.lower() == 'false':
                    default_params[param_name] = False
                elif param_value.lower() == 'none':
                    default_params[param_name] = None
                else:
                    try:
                        # Try as number
                        if '.' in param_value or 'e' in param_value.lower():
                            default_params[param_name] = float(param_value)
                        else:
                            default_params[param_name] = int(param_value)
                    except ValueError:
                        default_params[param_name] = param_value
        
        # If default_params is empty, assume using defaults
        use_defaults = len(default_params) == 0
        
        if name:  # Only add if we have a name
            algorithms.append(AlgorithmInfo(
                name=name,
                display_name=display_name,
                class_name=class_name,
                class_module=class_module,
                default_params=default_params,
                use_defaults=use_defaults,
            ))
    
    return algorithms


def _parse_experiment_groups_from_file(content: str) -> list[ExperimentGroupInfo]:
    """Parse experiment groups from settings.py content."""
    groups = []
    
    # Match add_experiment_group calls
    # Pattern is complex due to multi-line and various argument formats
    pattern = r'config\.add_experiment_group\s*\((.*?)\)'
    matches = re.findall(pattern, content, re.DOTALL)
    
    for match in matches:
        # Extract name
        name_match = re.search(r'name\s*=\s*["\']([^"\']+)["\']', match)
        name = name_match.group(1) if name_match else "Unknown"
        
        # Extract description
        desc_match = re.search(r'description\s*=\s*["\']([^"\']*)["\']', match)
        description = desc_match.group(1) if desc_match else ""
        
        # Extract datasets (can be a list)
        datasets_match = re.search(r'datasets\s*=\s*\[([^\]]*)\]', match)
        datasets = []
        if datasets_match:
            datasets_str = datasets_match.group(1)
            datasets = re.findall(r'["\']([^"\']+)["\']', datasets_str)
        
        # Extract algorithms (can be a list)
        algorithms_match = re.search(r'algorithms\s*=\s*\[([^\]]*)\]', match)
        algorithms = []
        if algorithms_match:
            algorithms_str = algorithms_match.group(1)
            algorithms = re.findall(r'["\']([^"\']+)["\']', algorithms_str)
        
        groups.append(ExperimentGroupInfo(
            name=name,
            description=description,
            datasets=datasets,
            algorithms=algorithms,
        ))
    
    return groups


@router.get("/experiments-data", response_model=ExperimentsDataResponse)
async def get_experiments_data(request: fastapi.Request):
    """Get all data needed for the experiments page.
    
    Returns:
    - Available datasets from the datasets folder
    - Configured algorithms from algorithms.py
    - Existing experiment groups from settings.py
    """
    settings = request.app.state.settings
    project_path = settings.project_path
    
    # Get datasets from datasets folder
    datasets: list[DatasetInfo] = []
    datasets_dir = project_path / "datasets"
    if datasets_dir.exists():
        for item in datasets_dir.iterdir():
            if item.is_file():
                suffix = item.suffix.lower()
                if suffix in ('.csv', '.xlsx', '.xls'):
                    datasets.append(DatasetInfo(
                        name=item.stem,  # filename without extension
                        filename=item.name,
                        file_type=suffix[1:],  # remove the dot
                    ))
                elif suffix in ('.sqlite', '.db', '.sqlite3'):
                    # For SQLite, we'd need to list tables - for now just list the file
                    datasets.append(DatasetInfo(
                        name=item.stem,
                        filename=item.name,
                        file_type="sqlite",
                    ))
    
    # Get algorithms from algorithms.py
    algorithms: list[AlgorithmInfo] = []
    algorithms_file = project_path / "algorithms.py"
    if algorithms_file.exists():
        try:
            content = algorithms_file.read_text()
            algorithms = _parse_algorithms_from_file(content)
        except Exception:
            pass
    
    # Get experiment groups from settings.py
    experiment_groups: list[ExperimentGroupInfo] = []
    settings_file = project_path / "settings.py"
    if settings_file.exists():
        try:
            content = settings_file.read_text()
            experiment_groups = _parse_experiment_groups_from_file(content)
        except Exception:
            pass
    
    return ExperimentsDataResponse(
        datasets=datasets,
        algorithms=algorithms,
        experiment_groups=experiment_groups,
    )


# ============================================================================
# Dataset File Parsing API
# ============================================================================

class FeatureInfo(pydantic.BaseModel):
    """Information about a single feature/column in a dataset."""
    name: str
    data_type: str  # "str", "int", or "float"
    categorical: bool = False  # User will mark this manually


class ParsedDatasetInfo(pydantic.BaseModel):
    """Parsed metadata from a dataset file."""
    file_name: str
    file_type: str  # "csv" or "xlsx"
    features: list[FeatureInfo]
    target_feature: str  # Last column name
    feature_count: int
    row_count: int  # Estimated row count


def _infer_dtype(values: list) -> str:
    """Infer data type from a sample of values.
    
    Returns "int", "float", or "str".
    """
    # Filter out None/empty values
    valid_values = [v for v in values if v is not None and str(v).strip() != ""]
    
    if not valid_values:
        return "str"
    
    # Try int first
    try:
        for v in valid_values:
            int(v)
        return "int"
    except (ValueError, TypeError):
        pass
    
    # Try float
    try:
        for v in valid_values:
            float(v)
        return "float"
    except (ValueError, TypeError):
        pass
    
    return "str"


@router.post("/parse-dataset-file", response_model=ParsedDatasetInfo)
async def parse_dataset_file(
    file: fastapi.UploadFile = fastapi.File(...),
):
    """Parse a dataset file and return metadata without loading full file into memory.
    
    Supports CSV and XLSX files. Reads only the first few rows to infer column types.
    The target feature is assumed to be the last column.
    """
    if not file.filename:
        raise fastapi.HTTPException(status_code=400, detail="No filename provided")
    
    file_ext = file.filename.split(".")[-1].lower() if "." in file.filename else ""
    
    if file_ext not in ("csv", "xlsx", "xls"):
        raise fastapi.HTTPException(
            status_code=400,
            detail=f"Unsupported file type: {file_ext}. Only CSV and XLSX are supported."
        )
    
    try:
        if file_ext == "csv":
            return await _parse_csv_file(file)
        else:  # xlsx or xls
            return await _parse_xlsx_file(file)
    except Exception as e:
        raise fastapi.HTTPException(
            status_code=500,
            detail=f"Failed to parse file: {str(e)}"
        )


async def _parse_csv_file(file: fastapi.UploadFile) -> ParsedDatasetInfo:
    """Parse CSV file metadata without loading entire file."""
    import csv
    import io
    
    # Read the file content - we need to do this to process it
    content = await file.read()
    
    # Now parse just the first few rows for column info
    text_content = content.decode('utf-8')
    reader = csv.reader(io.StringIO(text_content))
    
    # Get header
    try:
        header = next(reader)
    except StopIteration:
        raise fastapi.HTTPException(status_code=400, detail="Empty file")
    
    if not header:
        raise fastapi.HTTPException(status_code=400, detail="No columns found")
    
    # Read sample rows for type inference (up to 100 rows) and count total rows
    sample_rows = []
    row_count = 0
    for row in reader:
        row_count += 1
        if len(sample_rows) < 100:
            sample_rows.append(row)
    
    # Infer types for each column
    features: list[FeatureInfo] = []
    for col_idx, col_name in enumerate(header):
        col_values = [row[col_idx] for row in sample_rows if col_idx < len(row)]
        dtype = _infer_dtype(col_values)
        features.append(FeatureInfo(name=col_name.strip(), data_type=dtype))
    
    target_feature = header[-1].strip() if header else ""
    
    # feature_count excludes the target column
    return ParsedDatasetInfo(
        file_name=file.filename or "unknown.csv",
        file_type="csv",
        features=features,
        target_feature=target_feature,
        feature_count=len(features) - 1,
        row_count=row_count,
    )


async def _parse_xlsx_file(file: fastapi.UploadFile) -> ParsedDatasetInfo:
    """Parse XLSX file metadata without loading entire file."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise fastapi.HTTPException(
            status_code=500,
            detail="openpyxl is required for XLSX parsing. Install with: pip install openpyxl"
        )
    
    import io
    
    # Read file content
    content = await file.read()
    
    # Load workbook in read-only mode for memory efficiency
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    
    if ws is None:
        raise fastapi.HTTPException(status_code=400, detail="No active worksheet found")
    
    # Get header row (first row)
    header = []
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row is None:
        raise fastapi.HTTPException(status_code=400, detail="Empty file")
    
    header = [str(cell) if cell is not None else "" for cell in first_row]
    header = [h.strip() for h in header if h.strip()]
    
    if not header:
        raise fastapi.HTTPException(status_code=400, detail="No columns found")
    
    # Read sample rows for type inference (rows 2-101)
    sample_rows = []
    for row in ws.iter_rows(min_row=2, max_row=101, values_only=True):
        sample_rows.append(list(row))
    
    # Count total rows (this requires iterating but in read-only mode it's efficient)
    row_count = 0
    for _ in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
    
    wb.close()
    
    # Infer types for each column
    features: list[FeatureInfo] = []
    for col_idx, col_name in enumerate(header):
        col_values = [row[col_idx] for row in sample_rows if col_idx < len(row)]
        dtype = _infer_dtype(col_values)
        features.append(FeatureInfo(name=col_name, data_type=dtype))
    
    target_feature = header[-1] if header else ""
    
    # feature_count excludes the target column
    return ParsedDatasetInfo(
        file_name=file.filename or "unknown.xlsx",
        file_type="xlsx",
        features=features,
        target_feature=target_feature,
        feature_count=len(features) - 1,
        row_count=row_count,
    )


# ============================================================================
# Dataset Configuration Storage (project.json)
# ============================================================================

class StoredFeatureInfo(pydantic.BaseModel):
    """Stored feature information."""
    name: str
    data_type: str = "str"  # "str", "int", "float"
    categorical: bool = False


class StoredDataManagerConfig(pydantic.BaseModel):
    """Per-dataset data manager configuration."""
    test_size: float | None = None
    n_splits: int | None = None
    split_method: str | None = None
    group_column: str | None = None
    stratified: bool | None = None
    random_state: int | None = None


class StoredPreprocessorConfig(pydantic.BaseModel):
    """Preprocessor configuration for a dataset."""
    type: str  # "missing-data", "scaling", "encoding", "feature-selection"
    config: dict = {}


class StoredDatasetConfig(pydantic.BaseModel):
    """Complete stored configuration for a dataset.
    
    The ID is based on filename (and table name for SQLite).
    This ensures consistent IDs across sessions.
    """
    id: str  # filename or "filename:tablename" for sqlite
    file_name: str
    table_name: str | None = None
    file_type: str = "csv"  # "csv", "xlsx", "sqlite"
    target_feature: str = ""
    features_count: int = 0
    observations_count: int = 0
    features: list[StoredFeatureInfo] = []
    data_manager: StoredDataManagerConfig | None = None
    preprocessors: list[StoredPreprocessorConfig] = []


class StoredDatasetsResponse(pydantic.BaseModel):
    """Response containing stored datasets merged with file system."""
    datasets: list[StoredDatasetConfig]


class SaveDatasetsRequest(pydantic.BaseModel):
    """Request to save datasets configuration to project.json."""
    datasets: list[StoredDatasetConfig]


class SaveDatasetsResponse(pydantic.BaseModel):
    """Response for saving datasets."""
    success: bool
    saved_count: int


def _get_dataset_id(filename: str, table_name: str | None = None) -> str:
    """Generate a consistent dataset ID from filename and optional table name."""
    if table_name:
        return f"{filename}:{table_name}"
    return filename


def _list_dataset_files(project_path) -> list[dict]:
    """List dataset files in the datasets/ directory.
    
    Returns a list of dicts with file info.
    """
    import sqlite3
    
    datasets_dir = project_path / "datasets"
    files = []
    
    if not datasets_dir.exists():
        return files
    
    for item in datasets_dir.iterdir():
        if item.is_file():
            suffix = item.suffix.lower()
            if suffix in ('.csv', '.xlsx', '.xls'):
                files.append({
                    "id": item.name,
                    "file_name": item.name,
                    "table_name": None,
                    "file_type": "xlsx" if suffix in ('.xlsx', '.xls') else "csv",
                })
            elif suffix in ('.sqlite', '.db', '.sqlite3'):
                # For SQLite, list all tables
                try:
                    conn = sqlite3.connect(str(item))
                    cursor = conn.execute(
                        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
                    )
                    tables = [row[0] for row in cursor.fetchall()]
                    conn.close()
                    
                    for table in tables:
                        files.append({
                            "id": f"{item.name}:{table}",
                            "file_name": item.name,
                            "table_name": table,
                            "file_type": "sqlite",
                        })
                except Exception:
                    # If can't read database, just add the file
                    files.append({
                        "id": item.name,
                        "file_name": item.name,
                        "table_name": None,
                        "file_type": "sqlite",
                    })
    
    return files


@router.get("/datasets", response_model=StoredDatasetsResponse)
async def get_datasets(request: fastapi.Request):
    """Get datasets by merging file system with stored configuration.
    
    - Lists files in datasets/ directory
    - Loads stored configurations from project.json
    - Merges: if file exists, use stored config if available, otherwise create minimal entry
    - Ignores stored configs for files that no longer exist
    """
    settings = request.app.state.settings
    project_path = settings.project_path
    
    # Get files from file system
    file_entries = _list_dataset_files(project_path)
    
    # Load stored configs from project.json
    config_service = ProjectConfigService(project_path)
    config_data = config_service.read()
    stored_datasets = config_data.get("datasets", [])
    
    # Create a lookup by ID for stored configs
    stored_by_id = {d.get("id"): d for d in stored_datasets if d.get("id")}
    
    # Merge: use stored config if available, otherwise create minimal entry from file
    result_datasets: list[StoredDatasetConfig] = []
    
    for file_entry in file_entries:
        file_id = file_entry["id"]
        
        if file_id in stored_by_id:
            # Use stored config
            stored = stored_by_id[file_id]
            result_datasets.append(StoredDatasetConfig(
                id=file_id,
                file_name=stored.get("file_name", file_entry["file_name"]),
                table_name=stored.get("table_name", file_entry["table_name"]),
                file_type=stored.get("file_type", file_entry["file_type"]),
                target_feature=stored.get("target_feature", ""),
                features_count=stored.get("features_count", 0),
                observations_count=stored.get("observations_count", 0),
                features=[StoredFeatureInfo(**f) for f in stored.get("features", [])],
                data_manager=StoredDataManagerConfig(**stored["data_manager"]) if stored.get("data_manager") else None,
                preprocessors=[StoredPreprocessorConfig(**p) for p in stored.get("preprocessors", [])],
            ))
        else:
            # Create minimal entry from file info
            result_datasets.append(StoredDatasetConfig(
                id=file_id,
                file_name=file_entry["file_name"],
                table_name=file_entry["table_name"],
                file_type=file_entry["file_type"],
            ))
    
    return StoredDatasetsResponse(datasets=result_datasets)


@router.patch("/datasets", response_model=SaveDatasetsResponse)
async def save_datasets(
    request: fastapi.Request,
    data: SaveDatasetsRequest,
):
    """Save datasets configuration to project.json.
    
    This persists all dataset metadata, data manager configs, and preprocessor configs.
    """
    settings = request.app.state.settings
    project_path = settings.project_path
    
    config_service = ProjectConfigService(project_path)
    config_data = config_service.read()
    
    # Convert to dict format for JSON storage
    datasets_to_store = []
    for ds in data.datasets:
        ds_dict = {
            "id": ds.id,
            "file_name": ds.file_name,
            "table_name": ds.table_name,
            "file_type": ds.file_type,
            "target_feature": ds.target_feature,
            "features_count": ds.features_count,
            "observations_count": ds.observations_count,
            "features": [f.model_dump() for f in ds.features],
        }
        
        if ds.data_manager:
            ds_dict["data_manager"] = ds.data_manager.model_dump(exclude_none=True)
        
        if ds.preprocessors:
            ds_dict["preprocessors"] = [p.model_dump() for p in ds.preprocessors]
        
        datasets_to_store.append(ds_dict)
    
    config_data["datasets"] = datasets_to_store
    config_service.write(config_data)
    
    return SaveDatasetsResponse(
        success=True,
        saved_count=len(datasets_to_store),
    )
